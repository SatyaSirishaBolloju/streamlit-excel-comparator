import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
from fuzzywuzzy import process
import io

# Title and instructions for the app
st.title("📊 Excel Comparator Automation Tool")
st.markdown("""
Upload two Excel files:
- One **base file** where comparison blocks will be inserted
- One **source file** with reference sheets to pull data from

The tool uses **fuzzy matching** to match sheets by name and inserts a **12x8 block** with comparisons.
""")

# Upload the base Excel file (target workbook)
base_file = st.file_uploader("Upload Base Excel File", type=["xlsx"])

# Upload the source Excel file (reference workbook)
source_file = st.file_uploader("Upload Source Excel File", type=["xlsx"])

# Run logic only after both files are uploaded
if base_file and source_file:
    wb = openpyxl.load_workbook(base_file)       # Load the base workbook
    wb2 = openpyxl.load_workbook(source_file)    # Load the source workbook

    ws = wb.active                               # Use the active sheet from base workbook
    ws2_sheets = wb2.sheetnames                  # Get all sheet names from source workbook

    # Iterate through rows in base sheet (assuming headers in row 1)
    for i in range(2, ws.max_row + 1):
        shape = ws[f"B{i}"].value                # Get shape from column B
        weight_group = ws[f"F{i}"].value         # Get weight group from column F
        sheet_name = ws[f"M{i}"].value           # Get expected sheet name from column M

        # Skip rows with missing required data
        if not (shape and weight_group and sheet_name):
            continue

        # Find the closest matching sheet name using fuzzy string matching
        matched_sheet_name, score = process.extractOne(sheet_name, ws2_sheets)
        matched_sheet = wb2[matched_sheet_name]

        # Locate the cell in source sheet that matches the weight group
        match_cell = None
        for row in matched_sheet.iter_rows():
            for cell in row:
                if str(cell.value).strip().lower() == str(weight_group).strip().lower():
                    match_cell = cell
                    break
            if match_cell:
                break

        # Skip if weight group not found in source sheet
        if not match_cell:
            continue

        # Extract 12x8 block of data starting from the matched cell
        block = []
        for r in range(12):
            block_row = []
            for c in range(8):
                block_row.append(matched_sheet.cell(row=match_cell.row + r, column=match_cell.column + c).value)
            block.append(block_row)

        # Paste the block into the base sheet starting at column N (14)
        for r in range(12):
            for c in range(8):
                ws.cell(row=i + r, column=14 + c).value = block[r][c]

        # Apply border and formatting to the pasted block
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        for r in range(12):
            for c in range(8):
                ws.cell(row=i + r, column=14 + c).border = border

        # Add comparison formulas and apply formatting
        for c in range(8):
            base_cell = ws.cell(row=i + 12, column=14 + c)         # Row after block
            original_cell = ws.cell(row=i + 10, column=14 + c)     # Reference value
            current_cell = ws.cell(row=i + 8, column=14 + c)       # Value to compare

            base_cell.value = f"={original_cell.coordinate}-{current_cell.coordinate}"  # Difference formula
            base_cell.font = Font(bold=True)
            base_cell.border = border

            # Highlight differences with color (red = decrease, green = increase)
            try:
                diff = original_cell.value - current_cell.value
                if diff > 0:
                    base_cell.fill = green_fill
                elif diff < 0:
                    base_cell.fill = red_fill
            except Exception:
                pass  # Skip formatting if values aren't numbers

    # Save workbook to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)

    # Success message and download button
    st.success("✅ Excel updated successfully!")
    st.download_button(
        label="📥 Download Updated Excel",
        data=file_stream.getvalue(),
        file_name="updated_comparison.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
